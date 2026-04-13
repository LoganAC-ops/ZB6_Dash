#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SAP ECC -> S4 Output Accrual Comparison Tool
Mondelez International | Accenture
─────────────────────────────────────────────────────────────────
Usage:
  Double-click to open with file picker GUI, or:
  python compare_sap.py <ecc_file.xlsx> <s4_file.xlsx>
"""

import sys
import os
import re
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── Brand Colors ─────────────────────────────────────────────────────────────
BANNER_BG  = "1E0A4C"   # Deep Mondelez dark purple
HDR_BG     = "5A1F8C"   # Mondelez purple — column headers
SEC_BG     = "2D0E6A"   # Section dividers

MATCH_BG   = "E8F5E9"   # Soft green   — present in both
MISSING_BG = "FFF3CD"   # Soft amber   — missing in S4
EXTRA_BG   = "E3F2FD"   # Soft blue    — extra/new in S4
ROW_ALT    = "F9F7FE"   # Very light purple — alternating row
WHITE      = "FFFFFF"

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(size=10, bold=False, color="212121", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _border():
    s = Side(style="thin", color="D0C8E8")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(cell, value, bg=None, size=10, bold=False, color="212121",
         h="left", wrap=False, italic=False):
    cell.value = value
    if bg:
        cell.fill = _fill(bg)
    cell.font = _font(size=size, bold=bold, color=color, italic=italic)
    cell.alignment = _align(h=h, wrap=wrap)
    cell.border = _border()


# ─── Parsing ──────────────────────────────────────────────────────────────────

def parse_file(filepath):
    """
    Parse SAP XML/HTML output file into two flat lists:
      header_rows: one entry per HeaderFreeText (TextTypeCode + Text)
      line_rows:   one entry per ItemAmountsCharges within each LineItemInformation
    Namespace prefixes are stripped before parsing so tag names work regardless
    of which SAP system produced the file.
    """
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()

    # Strip XML namespace declarations and prefixes from tags
    content = re.sub(r'\s+xmlns(?::\w+)?="[^"]*"', '', content)
    content = re.sub(r'<([\w.-]+):([\w.-]+)', r'<\2', content)
    content = re.sub(r'</([\w.-]+):([\w.-]+)', r'</\2', content)

    root = ET.fromstring(content)

    def _text(node, tag):
        el = node.find(f".//{tag}")
        return el.text.strip() if el is not None and el.text else None

    doc_number  = _text(root, "DocumentNumber")
    header_rows = []
    line_rows   = []

    for hft in root.findall(".//HeaderFreeText"):
        text_type = _text(hft, "TextTypeCode")
        if text_type:
            header_rows.append({
                "field":   text_type,
                "value":   _text(hft, "Text") or "",
                "row_num": None,
            })

    for li in root.findall(".//LineItemInformation"):
        line_num     = _text(li, "LineItemNumber")
        material_num = _text(li, "MaterialNumber")
        material_desc= _text(li, "MaterialDescription")
        product_desc = _text(li, "ProductDescription")
        discount_amt = _text(li, "LineItemDiscountAmount")
        net_weight   = _text(li, "NetWeight")

        for iac in li.findall("ItemAmountsCharges"):
            if line_num:
                line_rows.append({
                    "line_num":      line_num,
                    "material_num":  material_num,
                    "material_desc": material_desc,
                    "product_desc":  product_desc,
                    "discount_amt":  discount_amt,
                    "charge_type":   _text(iac, "ChargeTypeCode"),
                    "description":   _text(iac, "Description"),
                    "amount":        _text(iac, "Amount"),
                    "net_weight":    net_weight,
                    "row_num":       None,
                })

    return header_rows, line_rows, doc_number


# ─── Date Validation ──────────────────────────────────────────────────────────

DATE_FIELDS = {"SODate", "DateAsoc", "From", "to"}

def check_dates(header_rows):
    """
    For header fields that should be YYYYMMDD, return a list of
    {field, value, source} dicts for any that fail validation.
    'source' is passed in by the caller so the UI can label ECC vs S4.
    """
    issues = []
    for r in header_rows:
        if r["field"] in DATE_FIELDS:
            val = r["value"]
            valid = bool(val and re.match(r"^\d{8}$", val))
            if valid:
                try:
                    datetime.strptime(val, "%Y%m%d")
                except ValueError:
                    valid = False
            if not valid:
                issues.append({"field": r["field"], "value": val or "(empty)"})
    return issues


# ─── Comparison Logic ─────────────────────────────────────────────────────────

def compare_headers(ecc_hdr, s4_hdr):
    """
    Compare all TextTypeCode fields across both files (flat, no grouping).
    ECC order first, then any S4-only fields appended at the end.
    Returns list of {field, ecc_value, s4_value, status}.
    """
    ecc_map = {r["field"]: r["value"] for r in ecc_hdr}
    s4_map  = {r["field"]: r["value"] for r in s4_hdr}

    ordered, seen = [], set()
    for r in ecc_hdr:
        if r["field"] not in seen:
            ordered.append(r["field"])
            seen.add(r["field"])
    for r in s4_hdr:
        if r["field"] not in seen:
            ordered.append(r["field"])
            seen.add(r["field"])

    results = []
    for field in ordered:
        in_ecc = field in ecc_map
        in_s4  = field in s4_map
        if in_ecc and in_s4:
            status = "match"
        else:
            status = "missing_in_s4" if in_ecc else "extra_in_s4"
        results.append({
            "field":     field,
            "ecc_value": ecc_map.get(field),
            "s4_value":  s4_map.get(field),
            "status":    status,
        })
    return results


def compare_line_items(ecc_lines, s4_lines):
    """
    Compare line item rows across both files (flat, no grouping by Serie).
    Match key: (line_num, charge_type).
    Returns list of {key, ecc, s4, status}.
    """
    def _key(r):
        return (str(r.get("line_num") or ""), str(r.get("charge_type") or ""))

    ecc_map, s4_map = {}, {}
    for r in ecc_lines:
        ecc_map.setdefault(_key(r), []).append(r)
    for r in s4_lines:
        s4_map.setdefault(_key(r), []).append(r)

    all_keys, seen = [], set()
    for r in ecc_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)
    for r in s4_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)

    results = []
    for k in all_keys:
        ecc_list = ecc_map.get(k, [])
        s4_list  = s4_map.get(k, [])
        for i in range(max(len(ecc_list), len(s4_list))):
            ecc_r = ecc_list[i] if i < len(ecc_list) else None
            s4_r  = s4_list[i]  if i < len(s4_list)  else None
            if ecc_r and s4_r:
                status = "match"
            else:
                status = "missing_in_s4" if ecc_r else "extra_in_s4"
            results.append({"key": k, "ecc": ecc_r, "s4": s4_r, "status": status})
    return results


# ─── Excel Report Helpers ─────────────────────────────────────────────────────

def _banner(ws, row, ecc_name, s4_name):
    """Write the top branded banner. Returns next row."""
    n = 9

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    _set(ws.cell(row, 1), "MONDELEZ INTERNATIONAL  |  ACCENTURE",
         bg=BANNER_BG, size=13, bold=True, color="FFFFFF", h="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    _set(ws.cell(row, 1), "E-Invoicing Comparison Report",
         bg=SEC_BG, size=11, bold=True, color="D8C8FF", h="center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    label = (f"ECC Baseline: {ecc_name}   <->   S4 New System: {s4_name}"
             f"   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    _set(ws.cell(row, 1), label, bg="F3EEF9", size=9, color="555555", h="center", italic=True)
    ws.row_dimensions[row].height = 16
    row += 1

    return row


def _legend(ws, row):
    ws.row_dimensions[row].height = 18
    ws.cell(row, 1).value = "KEY:"
    ws.cell(row, 1).font = _font(9, bold=True, color="444444")
    items = [
        ("  MATCH — values identical  ",           MATCH_BG,   "1B5E20"),
        ("  DIFFERENT — values changed  ",         "FFEBEE",   "B71C1C"),
        ("  MISSING IN S4 — in ECC only  ",        MISSING_BG, "7B4600"),
        ("  EXTRA IN S4 — not in ECC  ",           EXTRA_BG,   "0D47A1"),
    ]
    for i, (label, bg, fg) in enumerate(items, 2):
        c = ws.cell(row, i)
        _set(c, label, bg=bg, size=9, bold=True, color=fg, h="center")
    return row + 2


def _section_hdr(ws, row, title, n=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    _set(ws.cell(row, 1), f"  {title}", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
    for col in range(2, n + 1):
        ws.cell(row, col).fill = _fill(SEC_BG)
    ws.row_dimensions[row].height = 20
    return row + 1


def _col_hdrs(ws, row, headers, height=28):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.fill = _fill(HDR_BG)
        c.font = _font(9, bold=True, color="FFFFFF")
        c.alignment = _align(h="center", wrap=True)
        c.border = _border()
    ws.row_dimensions[row].height = height
    return row + 1


def _status_style(status):
    if status == "match":
        return MATCH_BG, "MATCH", "1B5E20"
    if status == "missing_in_s4":
        return MISSING_BG, "MISSING IN S4", "7B4600"
    if status == "extra_in_s4":
        return EXTRA_BG, "EXTRA IN S4", "0D47A1"
    return "FFEBEE", "DIFFERENT", "B71C1C"


# ─── Report Builder ───────────────────────────────────────────────────────────

def build_report(ecc_path, s4_path, output_path=None):
    ecc_name = os.path.basename(ecc_path)
    s4_name  = os.path.basename(s4_path)

    print(f"  Reading ECC : {ecc_name}")
    ecc_hdr, ecc_lines, ecc_docnum = parse_file(ecc_path)

    print(f"  Reading S4  : {s4_name}")
    s4_hdr,  s4_lines,  s4_docnum  = parse_file(s4_path)

    ecc_label = ecc_docnum or ecc_name
    s4_label  = s4_docnum  or s4_name

    print(f"  ECC header rows : {len(ecc_hdr)}   line rows : {len(ecc_lines)}")
    print(f"  S4  header rows : {len(s4_hdr)}   line rows : {len(s4_lines)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Column widths
    for i, w in enumerate([22, 30, 30, 18, 18, 14, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = _banner(ws, row, ecc_name, s4_name)
    row = _legend(ws, row)

    # ── SECTION 1: HEADER DETAILS ─────────────────────────────────────────────
    hdr_results  = compare_headers(ecc_hdr, s4_hdr)
    n_match      = sum(1 for r in hdr_results if r["status"] == "match")
    n_missing    = sum(1 for r in hdr_results if r["status"] == "missing_in_s4")
    n_extra      = sum(1 for r in hdr_results if r["status"] == "extra_in_s4")

    row = _section_hdr(ws, row,
        f"SECTION 1 — HEADER DETAILS  (Column J)   "
        f"| ECC: {len(ecc_hdr)} fields   S4: {len(s4_hdr)} fields   "
        f"Match: {n_match}   Missing in S4: {n_missing}   Extra in S4: {n_extra}",
        n=4)

    row = _col_hdrs(ws, row, [
        "FIELD  (TextTypeCode)",
        f"ECC Value\n({ecc_label})",
        f"S4 Value\n({s4_label})",
        "STATUS",
    ], height=24)

    if not hdr_results:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1, "No header data found in either file.").font = _font(9, italic=True, color="888888")
        row += 1
    else:
        for item in hdr_results:
            bg, status_text, fg = _status_style(item["status"])
            vals   = [item["field"],
                      item["ecc_value"] if item["ecc_value"] is not None else "—",
                      item["s4_value"]  if item["s4_value"]  is not None else "—",
                      status_text]
            aligns = ["left", "left", "left", "center"]
            bolds  = [False, False, False, True]
            for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
                c = ws.cell(row, i, v)
                c.fill      = _fill(bg)
                c.font      = _font(10, bold=b, color=(fg if i == 4 else "212121"))
                c.alignment = _align(h=ha)
                c.border    = _border()
            ws.row_dimensions[row].height = 17
            row += 1

    row += 1

    # ── SECTION 2: LINE ITEM / PRICING DETAILS ────────────────────────────────
    line_results = compare_line_items(ecc_lines, s4_lines)
    n_match   = sum(1 for r in line_results if r["status"] == "match")
    n_missing = sum(1 for r in line_results if r["status"] == "missing_in_s4")
    n_extra   = sum(1 for r in line_results if r["status"] == "extra_in_s4")

    row = _section_hdr(ws, row,
        f"SECTION 2 — LINE ITEM / PRICING DETAILS  (Cols Z, AA, AB, AC, AK, AM, AN)   "
        f"| ECC: {len(ecc_lines)} rows   S4: {len(s4_lines)} rows   "
        f"Match: {n_match}   Missing in S4: {n_missing}   Extra in S4: {n_extra}",
        n=9)

    row = _col_hdrs(ws, row, [
        "Line #",
        "Charge Type",
        f"ECC: MaterialNumber\n({ecc_label})",
        f"S4: MaterialNumber\n({s4_label})",
        f"ECC: Amount\n({ecc_label})",
        f"S4: Amount\n({s4_label})",
        f"ECC: Description\n({ecc_label})",
        f"S4: Description\n({s4_label})",
        "STATUS",
    ], height=30)

    if not line_results:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row, 1, "No line item data found in either file.").font = _font(9, italic=True, color="888888")
        row += 1
    else:
        for item in line_results:
            bg, status_text, fg = _status_style(item["status"])
            e = item["ecc"] or {}
            s = item["s4"]  or {}
            vals = [
                str(e.get("line_num")      or s.get("line_num")      or "—"),
                str(e.get("charge_type")   or s.get("charge_type")   or "—"),
                e.get("material_num")  or "—",
                s.get("material_num")  or "—",
                e.get("amount")        or "—",
                s.get("amount")        or "—",
                e.get("description")   or "—",
                s.get("description")   or "—",
                status_text,
            ]
            aligns = ["center","center","left","left","center","center","left","left","center"]
            bolds  = [False, True, False, False, False, False, False, False, True]
            for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
                c = ws.cell(row, i, v)
                c.fill      = _fill(bg)
                c.font      = _font(10, bold=b, color=(fg if i == 9 else "212121"))
                c.alignment = _align(h=ha)
                c.border    = _border()
            ws.row_dimensions[row].height = 17
            row += 1

    # ── Output ────────────────────────────────────────────────────────────────
    if output_path is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(ecc_path))
        output_path = os.path.join(out_dir, f"SAP_Comparison_{ts}.xlsx")

    wb.save(output_path)
    print(f"\n  Report saved : {output_path}")
    return output_path


# ─── Raw Data Export ─────────────────────────────────────────────────────────

def build_raw_export(ecc_path, s4_path, output_path=None):
    """
    Build a two-sheet Excel with the raw parsed data from each file.
    Sheet names are the document numbers (or filenames as fallback).
    """
    ecc_hdr, ecc_lines, ecc_docnum = parse_file(ecc_path)
    s4_hdr,  s4_lines,  s4_docnum  = parse_file(s4_path)

    ecc_label = ecc_docnum or os.path.basename(ecc_path)
    s4_label  = s4_docnum  or os.path.basename(s4_path)

    wb = Workbook()

    for label, hdr_rows, line_rows in [
        (ecc_label, ecc_hdr, ecc_lines),
        (s4_label,  s4_hdr,  s4_lines),
    ]:
        ws = wb.create_sheet(title=label[:31])  # Excel sheet name max 31 chars

        # ── Header fields section ─────────────────────────────────────────────
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _set(ws.cell(row, 1), "HEADER FIELDS", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
        ws.cell(row, 2).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for col, title in [(1, "TextTypeCode"), (2, "Value")]:
            c = ws.cell(row, col, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color="FFFFFF")
            c.alignment = _align(h="center")
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        for r in hdr_rows:
            _set(ws.cell(row, 1), r["field"],  h="left")
            _set(ws.cell(row, 2), r["value"],  h="left")
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1

        # ── Line items section ────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        _set(ws.cell(row, 1), "LINE ITEMS", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
        for col in range(2, 9):
            ws.cell(row, col).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        line_cols = ["Line #", "Charge Type", "Material Number", "Description",
                     "Amount", "Material Desc", "Product Desc", "Net Weight"]
        for i, title in enumerate(line_cols, 1):
            c = ws.cell(row, i, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color="FFFFFF")
            c.alignment = _align(h="center", wrap=True)
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        for r in line_rows:
            vals = [
                r.get("line_num")      or "",
                r.get("charge_type")   or "",
                r.get("material_num")  or "",
                r.get("description")   or "",
                r.get("amount")        or "",
                r.get("material_desc") or "",
                r.get("product_desc")  or "",
                r.get("net_weight")    or "",
            ]
            for i, v in enumerate(vals, 1):
                _set(ws.cell(row, i), v, h="left" if i not in (1, 5) else "center")
            ws.row_dimensions[row].height = 16
            row += 1

        # Column widths
        for i, w in enumerate([10, 14, 18, 30, 14, 24, 24, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if output_path is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(ecc_path))
        output_path = os.path.join(out_dir, f"SAP_RawData_{ts}.xlsx")

    wb.save(output_path)
    return output_path


# ─── Costa Rica Parsing ──────────────────────────────────────────────────────

def parse_file_cr(filepath):
    """
    Parse Costa Rica NotaCreditoElectronica XML.
    Returns (header_rows, line_rows, doc_number).
    """
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()

    content = re.sub(r'\s+xmlns(?::\w+)?="[^"]*"', '', content)
    content = re.sub(r'<([\w.-]+):([\w.-]+)', r'<\2', content)
    content = re.sub(r'</([\w.-]+):([\w.-]+)', r'</\2', content)

    root = ET.fromstring(content)

    def _t(node, tag):
        el = node.find(f".//{tag}")
        return el.text.strip() if el is not None and el.text else None

    def _n(node, tag):
        if node is None:
            return None
        el = node.find(f".//{tag}")
        return el.text.strip() if el is not None and el.text else None

    doc_number = _t(root, "NumeroConsecutivo")

    emisor   = root.find(".//Emisor")
    receptor = root.find(".//Receptor")
    resumen  = root.find(".//ResumenFactura")
    moneda   = root.find(".//CodigoTipoMoneda")
    medio    = resumen.find("MedioPago") if resumen is not None else None
    ref      = root.find(".//InformacionReferencia")

    header_rows = []

    # Root-level document fields
    for field, value in [
        ("Clave",                        _t(root, "Clave")),
        ("ProveedorSistemas",             _t(root, "ProveedorSistemas")),
        ("CodigoActividadEmisor",         _t(root, "CodigoActividadEmisor")),
        ("NumeroConsecutivo",             _t(root, "NumeroConsecutivo")),
        ("FechaEmision",                  _t(root, "FechaEmision")),
        ("CondicionVenta",                _t(root, "CondicionVenta")),
        ("PlazoCredito",                  _t(root, "PlazoCredito")),
    ]:
        header_rows.append({"field": field, "value": value or "", "row_num": None})

    # Emisor
    for field, value in [
        ("Emisor - Nombre",               _n(emisor, "Nombre")),
        ("Emisor - NombreComercial",      _n(emisor, "NombreComercial")),
        ("Emisor - Identificacion",       _n(emisor, "Numero")),
        ("Emisor - Provincia",            _n(emisor, "Provincia")),
        ("Emisor - Canton",               _n(emisor, "Canton")),
        ("Emisor - Distrito",             _n(emisor, "Distrito")),
        ("Emisor - OtrasSenas",           _n(emisor, "OtrasSenas")),
        ("Emisor - Telefono",             _n(emisor, "NumTelefono")),
        ("Emisor - CorreoElectronico",    _n(emisor, "CorreoElectronico")),
    ]:
        header_rows.append({"field": field, "value": value or "", "row_num": None})

    # Receptor
    for field, value in [
        ("Receptor - Nombre",             _n(receptor, "Nombre")),
        ("Receptor - NombreComercial",    _n(receptor, "NombreComercial")),
        ("Receptor - Identificacion",     _n(receptor, "Numero")),
        ("Receptor - Provincia",          _n(receptor, "Provincia")),
        ("Receptor - Canton",             _n(receptor, "Canton")),
        ("Receptor - Distrito",           _n(receptor, "Distrito")),
        ("Receptor - OtrasSenas",         _n(receptor, "OtrasSenas")),
        ("Receptor - Telefono",           _n(receptor, "NumTelefono")),
        ("Receptor - CorreoElectronico",  _n(receptor, "CorreoElectronico")),
    ]:
        header_rows.append({"field": field, "value": value or "", "row_num": None})

    # ResumenFactura
    for field, value in [
        ("CodigoMoneda",                  _n(moneda,  "CodigoMoneda")),
        ("TipoCambio",                    _n(moneda,  "TipoCambio")),
        ("TotalServGravados",             _n(resumen, "TotalServGravados")),
        ("TotalServExentos",              _n(resumen, "TotalServExentos")),
        ("TotalMercanciasGravadas",       _n(resumen, "TotalMercanciasGravadas")),
        ("TotalMercanciasExentas",        _n(resumen, "TotalMercanciasExentas")),
        ("TotalGravado",                  _n(resumen, "TotalGravado")),
        ("TotalExento",                   _n(resumen, "TotalExento")),
        ("TotalVenta",                    _n(resumen, "TotalVenta")),
        ("TotalDescuentos",               _n(resumen, "TotalDescuentos")),
        ("TotalVentaNeta",                _n(resumen, "TotalVentaNeta")),
        ("TotalImpuesto",                 _n(resumen, "TotalImpuesto")),
        ("MedioPago - TipoMedioPago",     _n(medio,   "TipoMedioPago")),
        ("MedioPago - TotalMedioPago",    _n(medio,   "TotalMedioPago")),
        ("TotalComprobante",              _n(resumen, "TotalComprobante")),
    ]:
        header_rows.append({"field": field, "value": value or "", "row_num": None})

    # InformacionReferencia — structure varies by file, only add if present
    if ref is not None:
        for tag, label in [
            ("TipoDocIR", "Referencia - TipoDocIR"),
            ("Codigo",    "Referencia - Codigo"),
            ("Numero",    "Referencia - Numero"),
            ("FechaEmisionIR", "Referencia - FechaEmisionIR"),
            ("Razon",     "Referencia - Razon"),
        ]:
            val = _n(ref, tag)
            if val is not None:
                header_rows.append({"field": label, "value": val, "row_num": None})

    # Line items
    line_rows = []
    for ld in root.findall(".//LineaDetalle"):
        codigo_interno = codigo_externo = None
        for cc in ld.findall("CodigoComercial"):
            tipo   = _t(cc, "Tipo")
            codigo = _t(cc, "Codigo")
            if tipo == "01":
                codigo_interno = codigo
            elif tipo == "03":
                codigo_externo = codigo
        impuesto = ld.find("Impuesto")
        line_rows.append({
            "line_num":               _t(ld, "NumeroLinea"),
            "partida_arancelaria":    _t(ld, "PartidaArancelaria"),
            "codigo_cabys":           _t(ld, "CodigoCABYS"),
            "codigo_interno":         codigo_interno,
            "codigo_externo":         codigo_externo,
            "detalle":                _t(ld, "Detalle"),
            "cantidad":               _t(ld, "Cantidad"),
            "unidad_medida":          _t(ld, "UnidadMedida"),
            "unidad_medida_comercial":_t(ld, "UnidadMedidaComercial"),
            "tipo_transaccion":       _t(ld, "TipoTransaccion"),
            "precio_unitario":        _t(ld, "PrecioUnitario"),
            "monto_total":            _t(ld, "MontoTotal"),
            "subtotal":               _t(ld, "SubTotal"),
            "base_imponible":         _t(ld, "BaseImponible"),
            "monto_total_linea":      _t(ld, "MontoTotalLinea"),
            "impuesto_codigo":        _n(impuesto, "Codigo")           if impuesto is not None else None,
            "impuesto_codigo_tarifa": _n(impuesto, "CodigoTarifaIVA") if impuesto is not None else None,
            "impuesto_tarifa":        _n(impuesto, "Tarifa")           if impuesto is not None else None,
            "impuesto_monto":         _n(impuesto, "Monto")            if impuesto is not None else None,
            "impuesto_asumido":       _t(ld, "ImpuestoAsumidoEmisorFabrica"),
            "impuesto_neto":          _t(ld, "ImpuestoNeto"),
        })

    return header_rows, line_rows, doc_number


def compare_cr_lines(prod_lines, test_lines):
    """
    Compare Costa Rica line items by NumeroLinea.
    Returns list of {key, prod, test, status}.
    """
    def _key(r):
        return str(r.get("line_num") or "")

    prod_map, test_map = {}, {}
    for r in prod_lines:
        prod_map.setdefault(_key(r), []).append(r)
    for r in test_lines:
        test_map.setdefault(_key(r), []).append(r)

    all_keys, seen = [], set()
    for r in prod_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)
    for r in test_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)

    results = []
    for k in all_keys:
        prod_list = prod_map.get(k, [])
        test_list = test_map.get(k, [])
        for i in range(max(len(prod_list), len(test_list))):
            prod_r = prod_list[i] if i < len(prod_list) else None
            test_r = test_list[i] if i < len(test_list) else None
            status = "match" if (prod_r and test_r) else ("missing_in_s4" if prod_r else "extra_in_s4")
            results.append({"key": k, "prod": prod_r, "test": test_r, "status": status})
    return results


def build_report_cr(prod_path, test_path, output_path=None):
    """Build Excel comparison report for Costa Rica."""
    prod_name = os.path.basename(prod_path)
    test_name = os.path.basename(test_path)

    prod_hdr, prod_lines, prod_docnum = parse_file_cr(prod_path)
    test_hdr, test_lines, test_docnum = parse_file_cr(test_path)

    prod_label = prod_docnum or prod_name
    test_label = test_docnum or test_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    for i, w in enumerate([28, 32, 32, 10, 10, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = _banner(ws, row, prod_name, test_name)
    row = _legend(ws, row)

    # ── Section 1: Headers ────────────────────────────────────────────────────
    hdr_results = compare_headers(prod_hdr, test_hdr)
    n_match   = sum(1 for r in hdr_results if r["status"] == "match")
    n_missing = sum(1 for r in hdr_results if r["status"] == "missing_in_s4")
    n_extra   = sum(1 for r in hdr_results if r["status"] == "extra_in_s4")

    row = _section_hdr(ws, row,
        f"SECTION 1 — HEADER DETAILS  "
        f"| Production: {len(prod_hdr)} fields   Testing: {len(test_hdr)} fields   "
        f"Match: {n_match}   Missing in Testing: {n_missing}   Extra in Testing: {n_extra}",
        n=4)

    row = _col_hdrs(ws, row, [
        "FIELD",
        f"Production Value\n({prod_label})",
        f"Testing Value\n({test_label})",
        "STATUS",
    ], height=24)

    for item in hdr_results:
        bg, status_text, fg = _status_style(item["status"])
        vals   = [item["field"],
                  item["ecc_value"] if item["ecc_value"] is not None else "—",
                  item["s4_value"]  if item["s4_value"]  is not None else "—",
                  status_text]
        aligns = ["left", "left", "left", "center"]
        bolds  = [False, False, False, True]
        for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
            c = ws.cell(row, i, v)
            c.fill      = _fill(bg)
            c.font      = _font(10, bold=b, color=(fg if i == 4 else "212121"))
            c.alignment = _align(h=ha)
            c.border    = _border()
        ws.row_dimensions[row].height = 17
        row += 1

    row += 1

    # ── Section 2: Line Items ─────────────────────────────────────────────────
    line_results = compare_cr_lines(prod_lines, test_lines)
    n_match   = sum(1 for r in line_results if r["status"] == "match")
    n_missing = sum(1 for r in line_results if r["status"] == "missing_in_s4")
    n_extra   = sum(1 for r in line_results if r["status"] == "extra_in_s4")

    row = _section_hdr(ws, row,
        f"SECTION 2 — LINE ITEMS  "
        f"| Production: {len(prod_lines)} rows   Testing: {len(test_lines)} rows   "
        f"Match: {n_match}   Missing in Testing: {n_missing}   Extra in Testing: {n_extra}",
        n=9)

    row = _col_hdrs(ws, row, [
        "Line #",
        f"Prod: Cod. Interno\n({prod_label})",
        f"Test: Cod. Interno\n({test_label})",
        "Detalle",
        f"Prod: PrecioUnitario\n({prod_label})",
        f"Test: PrecioUnitario\n({test_label})",
        f"Prod: MontoTotalLinea\n({prod_label})",
        f"Test: MontoTotalLinea\n({test_label})",
        "STATUS",
    ], height=30)

    for item in line_results:
        bg, status_text, fg = _status_style(item["status"])
        p = item["prod"] or {}
        t = item["test"] or {}
        vals = [
            str(p.get("line_num")          or t.get("line_num")          or "—"),
            p.get("codigo_interno")         or "—",
            t.get("codigo_interno")         or "—",
            str(p.get("detalle")           or t.get("detalle")           or "—"),
            p.get("precio_unitario")        or "—",
            t.get("precio_unitario")        or "—",
            p.get("monto_total_linea")      or "—",
            t.get("monto_total_linea")      or "—",
            status_text,
        ]
        aligns = ["center","left","left","left","center","center","center","center","center"]
        bolds  = [False]*8 + [True]
        for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
            c = ws.cell(row, i, v)
            c.fill      = _fill(bg)
            c.font      = _font(10, bold=b, color=(fg if i == 9 else "212121"))
            c.alignment = _align(h=ha)
            c.border    = _border()
        ws.row_dimensions[row].height = 17
        row += 1

    if output_path is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(prod_path))
        output_path = os.path.join(out_dir, f"CR_Comparison_{ts}.xlsx")

    wb.save(output_path)
    return output_path


def build_raw_export_cr(prod_path, test_path, output_path=None):
    """Build raw data Excel for Costa Rica (two sheets, one per file)."""
    prod_hdr, prod_lines, prod_docnum = parse_file_cr(prod_path)
    test_hdr, test_lines, test_docnum = parse_file_cr(test_path)

    prod_label = prod_docnum or os.path.basename(prod_path)
    test_label = test_docnum or os.path.basename(test_path)

    wb = Workbook()

    for label, hdr_rows, line_rows in [
        (prod_label, prod_hdr, prod_lines),
        (test_label, test_hdr, test_lines),
    ]:
        ws = wb.create_sheet(title=label[:31])
        row = 1

        # Header fields
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _set(ws.cell(row, 1), "HEADER FIELDS", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
        ws.cell(row, 2).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for col, title in [(1, "Field"), (2, "Value")]:
            c = ws.cell(row, col, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color="FFFFFF")
            c.alignment = _align(h="center")
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        for r in hdr_rows:
            _set(ws.cell(row, 1), r["field"], h="left")
            _set(ws.cell(row, 2), r["value"], h="left")
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1

        # Line items
        cr_line_cols = [
            "Line #", "PartidaArancelaria", "CodigoCABYS", "Cod. Interno", "Cod. Externo",
            "Detalle", "Cantidad", "UnidadMedida", "UnidadMedidaComercial", "TipoTransaccion",
            "PrecioUnitario", "MontoTotal", "SubTotal", "BaseImponible", "MontoTotalLinea",
            "Imp. Codigo", "Imp. CodTarifaIVA", "Imp. Tarifa", "Imp. Monto",
            "ImpuestoAsumido", "ImpuestoNeto",
        ]
        n_cr_line_cols = len(cr_line_cols)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cr_line_cols)
        _set(ws.cell(row, 1), "LINE ITEMS", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
        for col in range(2, n_cr_line_cols + 1):
            ws.cell(row, col).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for i, title in enumerate(cr_line_cols, 1):
            c = ws.cell(row, i, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color="FFFFFF")
            c.alignment = _align(h="center", wrap=True)
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        center_cols = {1, 7, 11, 12, 13, 14, 15, 18, 19, 20, 21}
        for r in line_rows:
            for i, v in enumerate([
                r.get("line_num")               or "",
                r.get("partida_arancelaria")     or "",
                r.get("codigo_cabys")            or "",
                r.get("codigo_interno")          or "",
                r.get("codigo_externo")          or "",
                r.get("detalle")                 or "",
                r.get("cantidad")                or "",
                r.get("unidad_medida")           or "",
                r.get("unidad_medida_comercial") or "",
                r.get("tipo_transaccion")        or "",
                r.get("precio_unitario")         or "",
                r.get("monto_total")             or "",
                r.get("subtotal")                or "",
                r.get("base_imponible")          or "",
                r.get("monto_total_linea")       or "",
                r.get("impuesto_codigo")         or "",
                r.get("impuesto_codigo_tarifa")  or "",
                r.get("impuesto_tarifa")         or "",
                r.get("impuesto_monto")          or "",
                r.get("impuesto_asumido")        or "",
                r.get("impuesto_neto")           or "",
            ], 1):
                _set(ws.cell(row, i), v, h="center" if i in center_cols else "left")
            ws.row_dimensions[row].height = 16
            row += 1

        for i, w in enumerate([6, 18, 18, 22, 16, 24, 10, 14, 18, 14,
                                16, 14, 14, 14, 16, 12, 16, 12, 12, 16, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if output_path is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(prod_path))
        output_path = os.path.join(out_dir, f"CR_RawData_{ts}.xlsx")

    wb.save(output_path)
    return output_path


# ─── Panama Parsing ──────────────────────────────────────────────────────────

def parse_file_pa(filepath):
    """
    Parse Panama MT_InvoiceRequest XML.
    Returns (header_rows, line_rows, doc_number).
    header_rows: [{field, value}] flat key-value pairs
    line_rows:   [{line_num, material_num, material_desc, ean, unit, quantity,
                   line_amount, tax_amount, tax_rate, gross_price, net_price,
                   discount_amount, taxable_amount, net_weight}]
    """
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()

    content = re.sub(r'\s+xmlns(?::\w+)?="[^"]*"', '', content)
    content = re.sub(r'<([\w.-]+):([\w.-]+)', r'<\2', content)
    content = re.sub(r'</([\w.-]+):([\w.-]+)', r'</\2', content)

    root = ET.fromstring(content)

    def _t(node, tag):
        el = node.find(f".//{tag}")
        return el.text.strip() if el is not None and el.text else None

    def _n(node, tag):
        if node is None:
            return None
        el = node.find(f".//{tag}")
        return el.text.strip() if el is not None and el.text else None

    doc_number = _t(root, "DocumentNumber")

    header_rows = []

    # Document-level fields
    for field, tag in [
        ("SAPSystem",         "SAPSystem"),
        ("EnvironmentID",     "EnvironmentID"),
        ("AreaID",            "AreaID"),
        ("ExternalNumber",    "ExternalNumber"),
        ("DocumentNumber",    "DocumentNumber"),
        ("CompanyCode",       "CompanyCode"),
        ("FiscalYear",        "FiscalYear"),
        ("DocumentType",      "DocumentType"),
        ("Serie",             "Serie"),
        ("Country",           "Country"),
        ("CreationDate",      "CreationDate"),
        ("CreationTime",      "CreationTime"),
        ("RefDocumentReason", "RefDocumentReason"),
    ]:
        header_rows.append({"field": field, "value": _t(root, tag) or "", "row_num": None})

    # Optional reference fields — only append if present in the XML
    doc_date = _t(root, "DocumentDate")
    if doc_date is not None:
        header_rows.append({"field": "DocumentDate", "value": doc_date, "row_num": None})

    # HeaderFreeText entries
    for hft in root.findall(".//HeaderFreeText"):
        text_type = _t(hft, "TextTypeCode")
        if text_type:
            header_rows.append({
                "field":   text_type,
                "value":   _t(hft, "Text") or "",
                "row_num": None,
            })

    # Party info (Emisor / Receptor)
    for party in root.findall(".//HeaderInformationParty"):
        role = _t(party, "PartyRoleCode") or "Unknown"
        header_rows.append({"field": f"{role} - PartyID",   "value": _t(party, "PartyID")   or "", "row_num": None})
        header_rows.append({"field": f"{role} - PartyType", "value": _t(party, "PartyType") or "", "row_num": None})
        header_rows.append({"field": f"{role} - Name",      "value": _t(party, "Name")      or "", "row_num": None})
        header_rows.append({"field": f"{role} - Address",   "value": _t(party, "Address")   or "", "row_num": None})
        add_data = party.find("HeaderInformationPartyAddData")
        if add_data is not None:
            email = _n(add_data, "EmailID")
            if email:
                header_rows.append({"field": f"{role} - EmailID", "value": email, "row_num": None})
            phone = _n(add_data, "PhoneNumber")
            if phone:
                header_rows.append({"field": f"{role} - Phone", "value": phone, "row_num": None})
            # OtherData entries (DataID / DataDetail pairs)
            for od in add_data.findall("OtherData"):
                data_id     = _n(od, "DataID")
                data_detail = _n(od, "DataDetail")
                if data_id and data_detail is not None:
                    header_rows.append({"field": f"{role} - {data_id}", "value": data_detail, "row_num": None})

    # Payment terms
    pt = root.find(".//HeaderInformationPaymentTerms")
    if pt is not None:
        for field, tag in [
            ("PaymentTermsTypeCode",      "PaymentTermsTypeCode"),
            ("PaymentDate",               "PaymentDate"),
            ("PaymentTermsDescription2",  "PaymentTermsDescription2"),
            ("PaymentTermsDescription3",  "PaymentTermsDescription3"),
        ]:
            header_rows.append({"field": field, "value": _n(pt, tag) or "", "row_num": None})

    # Total amounts — only append fields that are actually present in the XML
    ta = root.find(".//TotalAmounts")
    if ta is not None:
        for field, tag in [
            ("InvoiceAmount",            "InvoiceAmount"),
            ("SubTotal1",                "SubTotal1"),
            ("SubTotal2",                "SubTotal2"),
            ("SubTotal3",                "SubTotal3"),
            ("SubTotal4",                "SubTotal4"),
            ("TaxAmount",                "TaxAmount"),
            ("TotalForDiscount",         "TotalForDiscount"),
            ("TotalDiscountDescription", "TotalDiscountDescription"),
            ("TotalDiscountAmount",      "TotalDiscountAmount"),
        ]:
            val = _n(ta, tag)
            if val is not None:
                header_rows.append({"field": field, "value": val, "row_num": None})

    # Line items
    line_rows = []
    for li in root.findall(".//LineItemInformation"):
        pricing   = li.find("LineItemInformationQuantities/LineItemInformationPricingAndAmounts")
        discounts = li.find(".//LineItemPricingDiscounts")
        packaging = li.find(".//LineItemInformationPackagingDetails")
        line_rows.append({
            "line_num":        _t(li, "LineItemNumber"),
            "material_num":    _t(li, "MaterialNumber"),
            "material_desc":   _t(li, "MaterialDescription"),
            "ean":             _t(li, "ProductIDEAN"),
            "unit":            _t(li, "MeasureUnitCode"),
            "quantity":        _t(li, "InvoicedQuantity"),
            "line_amount":     _n(pricing, "LineItemAmount"),
            "tax_amount":      _n(pricing, "TaxAmount"),
            "taxable_amount":  _n(pricing, "TaxableAmount"),
            "tax_rate":        _n(pricing, "TaxRate"),
            "gross_price":     _n(pricing, "ProductGrossPrice"),
            "net_price":       _n(pricing, "ProductNetPrice"),
            "discount_amount": _n(discounts, "LineItemDiscountAmount") if discounts is not None else None,
            "net_weight":      _n(packaging, "NetWeight")              if packaging is not None else None,
            "gross_weight":    _n(packaging, "GrossWeight")            if packaging is not None else None,
        })

    return header_rows, line_rows, doc_number


def compare_pa_lines(prod_lines, test_lines):
    """
    Compare Panama line items by LineItemNumber.
    Returns list of {key, prod, test, status}.
    """
    def _key(r):
        return str(r.get("line_num") or "")

    prod_map, test_map = {}, {}
    for r in prod_lines:
        prod_map.setdefault(_key(r), []).append(r)
    for r in test_lines:
        test_map.setdefault(_key(r), []).append(r)

    all_keys, seen = [], set()
    for r in prod_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)
    for r in test_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)

    results = []
    for k in all_keys:
        prod_list = prod_map.get(k, [])
        test_list = test_map.get(k, [])
        for i in range(max(len(prod_list), len(test_list))):
            prod_r = prod_list[i] if i < len(prod_list) else None
            test_r = test_list[i] if i < len(test_list) else None
            status = "match" if (prod_r and test_r) else ("missing_in_s4" if prod_r else "extra_in_s4")
            results.append({"key": k, "prod": prod_r, "test": test_r, "status": status})
    return results


def build_report_pa(prod_path, test_path, output_path=None):
    """Build Excel comparison report for Panama."""
    prod_name = os.path.basename(prod_path)
    test_name = os.path.basename(test_path)

    prod_hdr, prod_lines, prod_docnum = parse_file_pa(prod_path)
    test_hdr, test_lines, test_docnum = parse_file_pa(test_path)

    prod_label = prod_docnum or prod_name
    test_label = test_docnum or test_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    for i, w in enumerate([28, 32, 32, 10, 10, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = _banner(ws, row, prod_name, test_name)
    row = _legend(ws, row)

    # ── Section 1: Headers ────────────────────────────────────────────────────
    hdr_results = compare_headers(prod_hdr, test_hdr)
    n_match   = sum(1 for r in hdr_results if r["status"] == "match")
    n_missing = sum(1 for r in hdr_results if r["status"] == "missing_in_s4")
    n_extra   = sum(1 for r in hdr_results if r["status"] == "extra_in_s4")

    row = _section_hdr(ws, row,
        f"SECTION 1 — HEADER DETAILS  "
        f"| Production: {len(prod_hdr)} fields   Testing: {len(test_hdr)} fields   "
        f"Match: {n_match}   Missing in Testing: {n_missing}   Extra in Testing: {n_extra}",
        n=4)

    row = _col_hdrs(ws, row, [
        "FIELD",
        f"Production Value\n({prod_label})",
        f"Testing Value\n({test_label})",
        "STATUS",
    ], height=24)

    for item in hdr_results:
        bg, status_text, fg = _status_style(item["status"])
        vals   = [item["field"],
                  item["ecc_value"] if item["ecc_value"] is not None else "—",
                  item["s4_value"]  if item["s4_value"]  is not None else "—",
                  status_text]
        aligns = ["left", "left", "left", "center"]
        bolds  = [False, False, False, True]
        for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
            c = ws.cell(row, i, v)
            c.fill      = _fill(bg)
            c.font      = _font(10, bold=b, color=(fg if i == 4 else "212121"))
            c.alignment = _align(h=ha)
            c.border    = _border()
        ws.row_dimensions[row].height = 17
        row += 1

    row += 1

    # ── Section 2: Line Items ─────────────────────────────────────────────────
    line_results = compare_pa_lines(prod_lines, test_lines)
    n_match   = sum(1 for r in line_results if r["status"] == "match")
    n_missing = sum(1 for r in line_results if r["status"] == "missing_in_s4")
    n_extra   = sum(1 for r in line_results if r["status"] == "extra_in_s4")

    row = _section_hdr(ws, row,
        f"SECTION 2 — LINE ITEMS  "
        f"| Production: {len(prod_lines)} rows   Testing: {len(test_lines)} rows   "
        f"Match: {n_match}   Missing in Testing: {n_missing}   Extra in Testing: {n_extra}",
        n=9)

    row = _col_hdrs(ws, row, [
        "Line #",
        f"Prod: MaterialNumber\n({prod_label})",
        f"Test: MaterialNumber\n({test_label})",
        "Material Desc",
        f"Prod: LineItemAmount\n({prod_label})",
        f"Test: LineItemAmount\n({test_label})",
        f"Prod: NetPrice\n({prod_label})",
        f"Test: NetPrice\n({test_label})",
        "STATUS",
    ], height=30)

    for item in line_results:
        bg, status_text, fg = _status_style(item["status"])
        p = item["prod"] or {}
        t = item["test"] or {}
        vals = [
            str(p.get("line_num")      or t.get("line_num")      or "—"),
            p.get("material_num")      or "—",
            t.get("material_num")      or "—",
            str(p.get("material_desc") or t.get("material_desc") or "—"),
            p.get("line_amount")       or "—",
            t.get("line_amount")       or "—",
            p.get("net_price")         or "—",
            t.get("net_price")         or "—",
            status_text,
        ]
        aligns = ["center","left","left","left","center","center","center","center","center"]
        bolds  = [False]*8 + [True]
        for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
            c = ws.cell(row, i, v)
            c.fill      = _fill(bg)
            c.font      = _font(10, bold=b, color=(fg if i == 9 else "212121"))
            c.alignment = _align(h=ha)
            c.border    = _border()
        ws.row_dimensions[row].height = 17
        row += 1

    if output_path is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(prod_path))
        output_path = os.path.join(out_dir, f"PA_Comparison_{ts}.xlsx")

    wb.save(output_path)
    return output_path


def build_raw_export_pa(prod_path, test_path, output_path=None):
    """Build raw data Excel for Panama (two sheets, one per file)."""
    prod_hdr, prod_lines, prod_docnum = parse_file_pa(prod_path)
    test_hdr, test_lines, test_docnum = parse_file_pa(test_path)

    prod_label = prod_docnum or os.path.basename(prod_path)
    test_label = test_docnum or os.path.basename(test_path)

    wb = Workbook()

    pa_line_cols = [
        "Line #", "MaterialNumber", "MaterialDesc", "EAN", "Unit", "Quantity",
        "LineItemAmount", "TaxAmount", "TaxableAmount", "TaxRate",
        "GrossPrice", "NetPrice", "DiscountAmount", "NetWeight", "GrossWeight",
    ]
    n_pa_line_cols = len(pa_line_cols)

    for label, hdr_rows, line_rows in [
        (prod_label, prod_hdr, prod_lines),
        (test_label, test_hdr, test_lines),
    ]:
        ws = wb.create_sheet(title=label[:31])
        row = 1

        # Header fields
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _set(ws.cell(row, 1), "HEADER FIELDS", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
        ws.cell(row, 2).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for col, title in [(1, "Field"), (2, "Value")]:
            c = ws.cell(row, col, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color="FFFFFF")
            c.alignment = _align(h="center")
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        for r in hdr_rows:
            _set(ws.cell(row, 1), r["field"], h="left")
            _set(ws.cell(row, 2), r["value"], h="left")
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1

        # Line items
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_pa_line_cols)
        _set(ws.cell(row, 1), "LINE ITEMS", bg=SEC_BG, size=10, bold=True, color="FFFFFF", h="left")
        for col in range(2, n_pa_line_cols + 1):
            ws.cell(row, col).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for i, title in enumerate(pa_line_cols, 1):
            c = ws.cell(row, i, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color="FFFFFF")
            c.alignment = _align(h="center", wrap=True)
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        center_cols = {1, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15}
        for r in line_rows:
            for i, v in enumerate([
                r.get("line_num")        or "",
                r.get("material_num")    or "",
                r.get("material_desc")   or "",
                r.get("ean")             or "",
                r.get("unit")            or "",
                r.get("quantity")        or "",
                r.get("line_amount")     or "",
                r.get("tax_amount")      or "",
                r.get("taxable_amount")  or "",
                r.get("tax_rate")        or "",
                r.get("gross_price")     or "",
                r.get("net_price")       or "",
                r.get("discount_amount") or "",
                r.get("net_weight")      or "",
                r.get("gross_weight")    or "",
            ], 1):
                _set(ws.cell(row, i), v, h="center" if i in center_cols else "left")
            ws.row_dimensions[row].height = 16
            row += 1

        for i, w in enumerate([6, 22, 28, 18, 10, 12, 16, 14, 14, 12, 14, 14, 16, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if output_path is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.dirname(os.path.abspath(prod_path))
        output_path = os.path.join(out_dir, f"PA_RawData_{ts}.xlsx")

    wb.save(output_path)
    return output_path


# ─── IDOC Parsing (Uruguay UY02 / Honduras HN02 / Venezuela VE02) ────────────

def parse_file_idoc(filepath):
    """
    Parse SAP IDOC HTML report (UY02, HN02, VE02 format).
    The table has 3 columns: Technical Name | Description | Value.
    Returns (header_rows, line_rows, doc_number).
    header_rows: [{field, value}] — all fields from E1EDK* segments
    line_rows:   [{line_num, quantity, unit, ean, net_amount, unit_price, total}]
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    def _clean(s):
        s = re.sub(r'&nbsp;', ' ', s)
        s = re.sub(r'&#x([0-9a-fA-F]+);', lambda m: chr(int(m.group(1), 16)), s)
        s = re.sub(r'&amp;', '&', s)
        s = re.sub(r'<[^>]+>', '', s)
        s = re.sub(r'\s+', ' ', s)
        return s.strip()

    # Extract ALL nobr contents (including empty — needed for correct triplet alignment)
    raw     = re.findall(r'<nobr[^>]*>(.*?)</nobr>', content, re.DOTALL)
    entries = [_clean(n) for n in raw]

    # First 3 entries = table header row (Technical Name, Description, Value)
    # Remaining: groups of 3 → (tech_name, description, value)
    rows = []
    i = 3
    while i + 2 < len(entries):
        tech = entries[i]
        val  = entries[i + 2]   # entries[i+1] is the description column (unused)
        rows.append((tech, val))
        i += 3

    SKIP = {'SEGNUM', 'EDIDC', 'EDIDD', 'STD', 'STDVRS', 'STDMES',
            'Technical Name', 'Description', 'Value'}

    header_rows = []
    line_rows   = []
    doc_number  = None

    in_line  = False
    cur_line = None

    for tech, val in rows:
        if not tech or tech in SKIP:
            continue

        if tech == 'SEGNAM':
            seg = val
            if seg == 'E1EDP01':
                # New line item — save previous
                if cur_line:
                    line_rows.append(cur_line)
                cur_line = {}
                in_line  = True
            elif re.match(r'(E1EDK|YOTC.*E1EDK)', seg):
                # Back to a header segment
                in_line = False
            # Other segments (E1EDP02/03/19, YOTC*_E1EDP01, etc.) keep current mode
            continue

        if in_line:
            if   tech == 'POSEX':           cur_line['line_num']   = val
            elif tech == 'MENGE':           cur_line['quantity']   = val
            elif tech == 'MENEE':           cur_line['unit']       = val
            # UY (YOTC10664_E1EDP01) extended fields
            elif tech == 'EAN11':           cur_line.setdefault('ean',        val)
            elif tech == 'NET_AMOUNT':      cur_line.setdefault('net_amount', val)
            elif tech == 'UNITPRICE':       cur_line.setdefault('unit_price', val)
            elif tech == 'MNTTOTAL':        cur_line.setdefault('total',      val)
            # HN / VE (YOTC_CRCM_E1EDP01) extended fields
            elif tech == 'YOTC_EAN11':      cur_line.setdefault('ean',        val)
            elif tech == 'YOTC_AMOUNT':     cur_line.setdefault('net_amount', val)
            elif tech == 'YOTC_UNIT_PRICE': cur_line.setdefault('unit_price', val)
            elif tech == 'YOTC_AMOUNT_TAX': cur_line.setdefault('total',      val)
            elif tech == 'ARKTX':           cur_line.setdefault('material_desc', val)
        else:
            if tech == 'BELNR' and not doc_number and val:
                doc_number = val
            if val:
                header_rows.append({'field': tech, 'value': val, 'row_num': None})

    if cur_line:
        line_rows.append(cur_line)

    return header_rows, line_rows, doc_number


def compare_idoc_lines(prod_lines, test_lines):
    """Compare IDOC line items by POSEX (line number)."""
    def _key(r):
        return str(r.get('line_num') or '')

    prod_map, test_map = {}, {}
    for r in prod_lines:
        prod_map.setdefault(_key(r), []).append(r)
    for r in test_lines:
        test_map.setdefault(_key(r), []).append(r)

    all_keys, seen = [], set()
    for r in prod_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)
    for r in test_lines:
        k = _key(r)
        if k not in seen:
            all_keys.append(k)
            seen.add(k)

    results = []
    for k in all_keys:
        pl = prod_map.get(k, [])
        tl = test_map.get(k, [])
        for i in range(max(len(pl), len(tl))):
            pr = pl[i] if i < len(pl) else None
            tr = tl[i] if i < len(tl) else None
            status = 'match' if (pr and tr) else ('missing_in_s4' if pr else 'extra_in_s4')
            results.append({'key': k, 'prod': pr, 'test': tr, 'status': status})
    return results


def build_report_idoc(prod_path, test_path, country_name, prefix, output_path=None):
    """Build Excel comparison report for IDOC countries (UY, HN, VE)."""
    prod_name = os.path.basename(prod_path)
    test_name = os.path.basename(test_path)

    prod_hdr, prod_lines, prod_docnum = parse_file_idoc(prod_path)
    test_hdr, test_lines, test_docnum = parse_file_idoc(test_path)

    prod_label = prod_docnum or prod_name
    test_label = test_docnum or test_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    for i, w in enumerate([28, 32, 32, 18, 18, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = _banner(ws, row, prod_name, test_name)
    row = _legend(ws, row)

    # ── Section 1: Headers ────────────────────────────────────────────────────
    hdr_results = compare_headers(prod_hdr, test_hdr)
    n_match   = sum(1 for r in hdr_results if r['status'] == 'match')
    n_missing = sum(1 for r in hdr_results if r['status'] == 'missing_in_s4')
    n_extra   = sum(1 for r in hdr_results if r['status'] == 'extra_in_s4')

    row = _section_hdr(ws, row,
        f"{country_name.upper()} — SECTION 1: HEADER DETAILS  "
        f"| Production: {len(prod_hdr)} fields   Testing: {len(test_hdr)} fields   "
        f"Match: {n_match}   Missing in Testing: {n_missing}   Extra in Testing: {n_extra}",
        n=4)

    row = _col_hdrs(ws, row, [
        "FIELD (Technical Name)",
        f"Production Value\n({prod_label})",
        f"Testing Value\n({test_label})",
        "STATUS",
    ], height=24)

    for item in hdr_results:
        bg, status_text, fg = _status_style(item['status'])
        vals   = [item['field'],
                  item['ecc_value'] if item['ecc_value'] is not None else '—',
                  item['s4_value']  if item['s4_value']  is not None else '—',
                  status_text]
        aligns = ['left', 'left', 'left', 'center']
        bolds  = [False, False, False, True]
        for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
            c = ws.cell(row, i, v)
            c.fill      = _fill(bg)
            c.font      = _font(10, bold=b, color=(fg if i == 4 else '212121'))
            c.alignment = _align(h=ha)
            c.border    = _border()
        ws.row_dimensions[row].height = 17
        row += 1

    row += 1

    # ── Section 2: Line Items ─────────────────────────────────────────────────
    line_results = compare_idoc_lines(prod_lines, test_lines)
    n_match   = sum(1 for r in line_results if r['status'] == 'match')
    n_missing = sum(1 for r in line_results if r['status'] == 'missing_in_s4')
    n_extra   = sum(1 for r in line_results if r['status'] == 'extra_in_s4')

    row = _section_hdr(ws, row,
        f"{country_name.upper()} — SECTION 2: LINE ITEMS  "
        f"| Production: {len(prod_lines)} rows   Testing: {len(test_lines)} rows   "
        f"Match: {n_match}   Missing in Testing: {n_missing}   Extra in Testing: {n_extra}",
        n=9)

    row = _col_hdrs(ws, row, [
        "Line #",
        f"Prod: EAN/Material\n({prod_label})",
        f"Test: EAN/Material\n({test_label})",
        "Qty",
        f"Prod: Net Amount\n({prod_label})",
        f"Test: Net Amount\n({test_label})",
        f"Prod: Unit Price\n({prod_label})",
        f"Test: Unit Price\n({test_label})",
        "STATUS",
    ], height=30)

    for item in line_results:
        bg, status_text, fg = _status_style(item['status'])
        p = item['prod'] or {}
        t = item['test'] or {}
        vals = [
            str(p.get('line_num')    or t.get('line_num')    or '—'),
            p.get('ean')             or '—',
            t.get('ean')             or '—',
            str(p.get('quantity')    or t.get('quantity')    or '—'),
            p.get('net_amount')      or '—',
            t.get('net_amount')      or '—',
            p.get('unit_price')      or '—',
            t.get('unit_price')      or '—',
            status_text,
        ]
        aligns = ['center', 'left', 'left', 'center', 'center', 'center', 'center', 'center', 'center']
        bolds  = [False] * 8 + [True]
        for i, (v, ha, b) in enumerate(zip(vals, aligns, bolds), 1):
            c = ws.cell(row, i, v)
            c.fill      = _fill(bg)
            c.font      = _font(10, bold=b, color=(fg if i == 9 else '212121'))
            c.alignment = _align(h=ha)
            c.border    = _border()
        ws.row_dimensions[row].height = 17
        row += 1

    if output_path is None:
        ts          = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_dir     = os.path.dirname(os.path.abspath(prod_path))
        output_path = os.path.join(out_dir, f'{prefix}_Comparison_{ts}.xlsx')

    wb.save(output_path)
    return output_path


def build_raw_export_idoc(prod_path, test_path, country_name, prefix, output_path=None):
    """Build raw data Excel for IDOC countries (UY, HN, VE)."""
    prod_hdr, prod_lines, prod_docnum = parse_file_idoc(prod_path)
    test_hdr, test_lines, test_docnum = parse_file_idoc(test_path)

    prod_label = prod_docnum or os.path.basename(prod_path)
    test_label = test_docnum or os.path.basename(test_path)

    idoc_line_cols = [
        "Line #", "EAN/Material", "Quantity", "Unit",
        "Net Amount", "Unit Price", "Total", "Material Desc",
    ]
    n_cols = len(idoc_line_cols)

    wb = Workbook()

    for label, hdr_rows, line_rows in [
        (prod_label, prod_hdr, prod_lines),
        (test_label, test_hdr, test_lines),
    ]:
        ws  = wb.create_sheet(title=label[:31])
        row = 1

        # Header fields
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _set(ws.cell(row, 1), 'HEADER FIELDS', bg=SEC_BG, size=10, bold=True, color='FFFFFF', h='left')
        ws.cell(row, 2).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for col, title in [(1, 'Technical Name'), (2, 'Value')]:
            c = ws.cell(row, col, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color='FFFFFF')
            c.alignment = _align(h='center')
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        for r in hdr_rows:
            _set(ws.cell(row, 1), r['field'], h='left')
            _set(ws.cell(row, 2), r['value'], h='left')
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1

        # Line items
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        _set(ws.cell(row, 1), 'LINE ITEMS', bg=SEC_BG, size=10, bold=True, color='FFFFFF', h='left')
        for col in range(2, n_cols + 1):
            ws.cell(row, col).fill = _fill(SEC_BG)
        ws.row_dimensions[row].height = 20
        row += 1

        for i, title in enumerate(idoc_line_cols, 1):
            c = ws.cell(row, i, title)
            c.fill = _fill(HDR_BG)
            c.font = _font(9, bold=True, color='FFFFFF')
            c.alignment = _align(h='center', wrap=True)
            c.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        center_cols = {1, 3, 5, 6, 7}
        for r in line_rows:
            for i, v in enumerate([
                r.get('line_num')      or '',
                r.get('ean')          or '',
                r.get('quantity')     or '',
                r.get('unit')         or '',
                r.get('net_amount')   or '',
                r.get('unit_price')   or '',
                r.get('total')        or '',
                r.get('material_desc') or '',
            ], 1):
                _set(ws.cell(row, i), v, h='center' if i in center_cols else 'left')
            ws.row_dimensions[row].height = 16
            row += 1

        for i, w in enumerate([8, 22, 10, 8, 16, 16, 16, 28], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    if output_path is None:
        ts          = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_dir     = os.path.dirname(os.path.abspath(prod_path))
        output_path = os.path.join(out_dir, f'{prefix}_RawData_{ts}.xlsx')

    wb.save(output_path)
    return output_path


# ─── Entry Point ──────────────────────────────────────────────────────────────

COUNTRY_BUILDERS = {
    "AR": build_report,
    "CR": build_report_cr,
    "PA": build_report_pa,
    "UY": lambda p, t, o=None: build_report_idoc(p, t, "Uruguay",   "UY", o),
    "HN": lambda p, t, o=None: build_report_idoc(p, t, "Honduras",  "HN", o),
    "VE": lambda p, t, o=None: build_report_idoc(p, t, "Venezuela", "VE", o),
}

COUNTRY_LABELS = {
    "AR": "Argentina",
    "CR": "Costa Rica",
    "PA": "Panama",
    "UY": "Uruguay (IDOC)",
    "HN": "Honduras (IDOC)",
    "VE": "Venezuela (IDOC)",
}


def _pick_country_gui(parent):
    """Show a simple radio-button dialog and return the chosen country code."""
    import tkinter as tk

    result = {"code": None}

    dlg = tk.Toplevel(parent)
    dlg.title("Select Country")
    dlg.attributes("-topmost", True)
    dlg.resizable(False, False)

    tk.Label(dlg, text="Select the country for this comparison:",
             font=("Calibri", 11, "bold"), pady=10).pack(padx=20)

    choice = tk.StringVar(value="AR")
    for code, label in COUNTRY_LABELS.items():
        tk.Radiobutton(dlg, text=f"{label} ({code})", variable=choice,
                       value=code, font=("Calibri", 10)).pack(anchor="w", padx=30)

    def _ok():
        result["code"] = choice.get()
        dlg.destroy()

    tk.Button(dlg, text="OK", command=_ok, width=10,
              font=("Calibri", 10, "bold")).pack(pady=12)

    dlg.grab_set()
    parent.wait_window(dlg)
    return result["code"]


def main():
    ecc_path = s4_path = country = None

    if len(sys.argv) == 4:
        country  = sys.argv[1].upper()
        ecc_path = sys.argv[2]
        s4_path  = sys.argv[3]
    elif len(sys.argv) == 3:
        ecc_path = sys.argv[1]
        s4_path  = sys.argv[2]
        country  = "AR"
    else:
        try:
            import tkinter as tk
            from tkinter import filedialog, messagebox

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            country = _pick_country_gui(root)
            if not country:
                root.destroy()
                return

            country_label = COUNTRY_LABELS.get(country, country)

            messagebox.showinfo(
                "SAP Comparison Tool — Mondelez | Accenture",
                f"Step 1 of 2 [{country_label}]: Select the Production (old system) XML file.",
                parent=root,
            )
            ecc_path = filedialog.askopenfilename(
                title=f"Select Production XML File [{country_label}]",
                filetypes=[("XML/HTML Files", "*.xml *.html"), ("All Files", "*.*")],
                parent=root,
            )
            if not ecc_path:
                messagebox.showwarning("Cancelled", "No Production file selected.", parent=root)
                root.destroy()
                return

            messagebox.showinfo(
                "SAP Comparison Tool — Mondelez | Accenture",
                f"Step 2 of 2 [{country_label}]: Select the Testing (new system) XML file.",
                parent=root,
            )
            s4_path = filedialog.askopenfilename(
                title=f"Select Testing XML File [{country_label}]",
                filetypes=[("XML/HTML Files", "*.xml *.html"), ("All Files", "*.*")],
                parent=root,
            )
            if not s4_path:
                messagebox.showwarning("Cancelled", "No Testing file selected.", parent=root)
                root.destroy()
                return

            root.destroy()
        except Exception as e:
            print(f"GUI unavailable: {e}")
            print("Usage: python compare_sap.py [AR|CR|PA] <prod_file.xml> <test_file.xml>")
            return

    if country not in COUNTRY_BUILDERS:
        print(f"Unknown country code '{country}'. Valid options: {', '.join(COUNTRY_BUILDERS)}")
        sys.exit(1)

    print(f"\nSAP Comparison Tool  |  Mondelez | Accenture  |  {COUNTRY_LABELS[country]}")
    print("=" * 60)
    try:
        output = COUNTRY_BUILDERS[country](ecc_path, s4_path)
        try:
            import subprocess
            subprocess.Popen(["start", "", output], shell=True)
        except Exception:
            pass
        print("\nDone.")
    except Exception as err:
        import traceback
        print(f"\nERROR: {err}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
